"""
================================================
ESP32 Smart Pump Controller v10 — Excel Logger
Includes Voltage @ Pump ON and Pump OFF columns
================================================
INSTALL ONCE:
    pip install pyserial openpyxl

CHANGE COM_PORT below then run:
    python excel_logger_v3.py
================================================
"""

import serial
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time
import os

# ================================================
# SETTINGS — CHANGE COM_PORT TO YOUR PORT
# ================================================
COM_PORT   = "COM3"              # ← CHANGE THIS
BAUD_RATE  = 115200
EXCEL_FILE = "sensor_data_v10.xlsx"
SAVE_EVERY = 5
# ================================================

GREEN_FILL   = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
YELLOW_FILL  = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
HEADER_FILL  = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
SOLAR_FILL   = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
VOLT_FILL    = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
THIN         = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

HEADERS = [
    "No.", "Date", "Time",
    "Temp (°C)", "Humidity (%)",
    "pH", "Soil (%)", "Air Quality",
    "N (mg/kg)", "P (mg/kg)", "K (mg/kg)",
    "Pump", "Solar V", "Solar mA", "Solar mW",
    "V @ Pump ON", "V @ Pump OFF"      # ← NEW
]

COL_WIDTHS = [5, 12, 10, 11, 13, 8, 9, 13, 12, 12, 12, 8, 10, 11, 11, 13, 14]


def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sensor Data"

    # Title
    last_col = openpyxl.utils.get_column_letter(len(HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "ESP32 Smart Pump Controller v10 — Sensor & Voltage Log"
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value     = "Jamia Millia Islamia University | Voltage Tracking @ Pump ON/OFF | ZTS-3002 NPK | INA219 Solar"
    s.font      = Font(italic=True, size=10, color="555555")
    s.alignment = CENTER
    ws.row_dimensions[2].height = 16

    # Headers row 3
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.alignment = CENTER
        cell.border    = THIN
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        # Colour header by group
        if col in [13, 14, 15]:
            cell.fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
            cell.font = HEADER_FONT
        elif col in [16, 17]:
            cell.fill = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
            cell.font = HEADER_FONT
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"
    return wb, ws


def load_or_create():
    if os.path.exists(EXCEL_FILE):
        wb  = openpyxl.load_workbook(EXCEL_FILE)
        ws  = wb.active
        cnt = max(0, ws.max_row - 3)
        print(f"  Loaded existing: {cnt} readings already saved")
        return wb, ws, cnt
    else:
        wb, ws = create_workbook()
        print(f"  Created new file: {EXCEL_FILE}")
        return wb, ws, 0


def parse_line(line):
    """
    DATA,temp,hum,ph,soil,air,N,P,K,pump,
         solarV,solarmA,solarmW,V@PumpON,V@PumpOFF
    Total 15 fields after DATA = 15 parts[1:]
    """
    parts = line.strip().split(",")
    if len(parts) != 15 or parts[0] != "DATA":
        return None
    try:
        return {
            "temp":    parts[1],
            "hum":     parts[2],
            "ph":      parts[3],
            "soil":    parts[4],
            "air":     parts[5],
            "N":       parts[6],
            "P":       parts[7],
            "K":       parts[8],
            "pump":    parts[9].strip(),
            "solarV":  parts[10],
            "solarmA": parts[11],
            "solarmW": parts[12],
            "vOn":     parts[13],
            "vOff":    parts[14].strip()
        }
    except Exception:
        return None


def append_row(ws, data, no):
    now  = datetime.now()
    row  = [
        no,
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        data["temp"],    data["hum"],
        data["ph"],      data["soil"],    data["air"],
        data["N"],       data["P"],       data["K"],
        data["pump"],
        data["solarV"],  data["solarmA"], data["solarmW"],
        data["vOn"],     data["vOff"]
    ]
    ws.append(row)
    lr = ws.max_row

    # Row colour based on pump/soil
    pump = data["pump"]
    try:    soil = int(data["soil"])
    except: soil = 50

    if pump == "ON":    row_fill = RED_FILL
    elif soil >= 60:    row_fill = GREEN_FILL
    else:               row_fill = YELLOW_FILL

    for col in range(1, len(HEADERS) + 1):
        cell           = ws.cell(row=lr, column=col)
        cell.alignment = CENTER
        cell.border    = THIN
        if col in [13, 14, 15]:
            cell.fill = SOLAR_FILL
        elif col in [16, 17]:
            cell.fill = VOLT_FILL     # blue tint for voltage columns
        else:
            cell.fill = row_fill

    # Bold pump ON
    if pump == "ON":
        ws.cell(row=lr, column=12).font = Font(bold=True, color="C62828")


def main():
    print()
    print("=" * 58)
    print("   ESP32 Pump Controller v10 — Excel Logger")
    print("=" * 58)
    print(f"  Port : {COM_PORT}  |  File: {EXCEL_FILE}")
    print("=" * 58)

    wb, ws, existing = load_or_create()
    reading_no = existing
    unsaved    = 0

    print("\n  Close Arduino Serial Monitor first!")
    print("  Connecting...\n")

    try:
        ser = serial.Serial(COM_PORT, BAUD_RATE, timeout=3)
        time.sleep(2)
        print(f"  Connected to {COM_PORT}")
        print(f"\n  {'#':>4}  {'Temp':>6} {'Soil':>5} {'Pump':<5} "
              f"{'SolarV':>7} {'V@ON':>7} {'V@OFF':>8}")
        print("  " + "-" * 52)

        while True:
            try:
                raw = ser.readline().decode("utf-8", errors="ignore").strip()
                if not raw.startswith("DATA,"):
                    continue

                data = parse_line(raw)
                if data is None:
                    continue

                reading_no += 1
                unsaved    += 1
                append_row(ws, data, reading_no)

                print(f"  {reading_no:>4}  "
                      f"{data['temp']:>6}°C "
                      f"{data['soil']:>4}%  "
                      f"{data['pump']:<5} "
                      f"{data['solarV']:>6}V "
                      f"ON:{data['vOn']:>5}V "
                      f"OFF:{data['vOff']:>5}V")

                if unsaved >= SAVE_EVERY:
                    wb.save(EXCEL_FILE)
                    unsaved = 0
                    print(f"\n  ✓ Saved — {reading_no} readings → {EXCEL_FILE}\n")

            except UnicodeDecodeError:
                continue

    except serial.SerialException as e:
        print(f"\n  ERROR: {e}")
        print(f"  → Check {COM_PORT} is correct")
        print(f"  → Close Arduino Serial Monitor first")

    except KeyboardInterrupt:
        print("\n\n  Stopped by user.")

    finally:
        try:
            wb.save(EXCEL_FILE)
            print(f"  ✓ Final save: {reading_no} readings → {EXCEL_FILE}")
        except Exception as ex:
            print(f"  Save error: {ex}")
        try:
            ser.close()
        except Exception:
            pass


if __name__ == "__main__":
    main()